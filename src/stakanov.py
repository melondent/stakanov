import os
import pandas as pd
from os.path import join, getsize, getctime, getmtime
from tqdm import tqdm
from datetime import datetime
from PyPDF2 import PdfReader
from PIL import Image
from docx import Document
from openpyxl import load_workbook
from mutagen.mp3 import MP3
import cv2
import humanize
import tkinter as tk
import logging

class Logger:
    def __init__(self, log_file='stakanov.log'):
        logging.basicConfig(
            filename=log_file,
            format='%(asctime)s - %(levelname)s - %(message)s',
            datefmt='%Y-%m-%d %H:%M:%S'
        )
        self.logger = logging.getLogger()

    def log_exception(self, message):
        self.logger.exception(message)

class Guide:
    def __init__(self, path):
        self.path = path
        
    def files(self):
        for root, dirs, files in os.walk(self.path):
            for f in files:
                yield join(root, f)
                
class Indiana:
    def  __init__(self, path, output_file):
        self.logger = Logger()
        self.guide = Guide(path)
        self.path = path
        self.researchers = [
            GeneralResearcher(), PDFResearcher(), ImageResearcher(), 
            DocResearcher(), ExcelResearcher(), AudioResearcher(), 
            VideoResearcher()
        ]
        self.output_file = output_file
        self.data = []
        self.csv_saver = Saver(output_file)
        self.viewer = Displayer(output_file)

    def find_loot(self, progress_callback=None):
        total_files = sum(1 for _ in self.guide.files())
        progress_callback(total_files)
        
        for file in tqdm(self.guide.files(), total=total_files):
            file_info = {}
            for researcher in self.researchers:
                try:
                    file_info.update(researcher.get_info(file))
                except Exception as e:
                    self.logger.log_exception(f"Ошибка при обработке файла {file}: {str(e)}")
            if file_info:
                file_info['file_path'] = file
                self.data.append(file_info)
                
            if progress_callback: 
                progress_callback(len(self.data))
                
    def save_results(self):
        self.csv_saver.save_scv(self.data)
    
    def display_results(self):
        self.viewer.display()
        

class GeneralResearcher:
    def get_info(self, file):
        size = getsize(file)
        name = os.path.splitext(os.path.basename(file))[0]
        extension = os.path.splitext(file)[1].replace('.', '').upper()
        return {
            'name': name,
            'extension': extension,
            'size': size,
            'created': datetime.fromtimestamp(getctime(file)).strftime('%Y-%m-%d %H:%M:%S'),
            'modified': datetime.fromtimestamp(getmtime(file)).strftime('%Y-%m-%d %H:%M:%S')
        }
        
    
class ImageResearcher:
    def get_info(self, file):
        if not file.lower().endswith(('.png', '.jpg', '.jpeg', '.gif')):
            return {}
        else:
            img = cv2.imread(file)
            height, width = img.shape[:2]
            img_pil = Image.open(file)
            dpi = img_pil.info.get('dpi', (None, None))
            if dpi != (None, None):
                dpi = (round(dpi[0]), round(dpi[1]))
            exif_data = img_pil._getexif()
            
            return {
                'width': width,
                'height': height,
                'dpi': dpi if dpi != (None, None) else 'No DPI data',
                'exif': exif_data if exif_data else 'No EXIF data'
            }
    
class DocResearcher:
    
    DOC_PAGE_SIZES = {
        'A0': (841, 1189),
        'A1': (594, 841),
        'A2': (420, 594),
        'A3': (297, 420),
        'A4': (210, 297),
        'A5': (148, 210),
        'A6': (105, 148)
    }
    
    def get_format(self, width, height):
        for format, (w, h) in self.DOC_PAGE_SIZES.items():
            if abs(width - w) < 10 and abs(height - h) < 10:
                return format
        return f'{int(width)}x{int(height)}'
    
    def get_orientation(self, section):
        return 'landscape' if section.orientation == 1 else 'portrait'
    
    def get_info(self, file):
        if not file.lower().endswith('.docx'):
            return {}
        else:
            doc = Document(file)
            core_properties = doc.core_properties
            section = doc.sections[0]
            return {
                'author': core_properties.author,
                'created': core_properties.created.strftime('%Y-%m-%d %H:%M:%S') if core_properties.created else GeneralResearcher().get_info(file)['created'],
                'modified': core_properties.modified.strftime('%Y-%m-%d %H:%M:%S') if core_properties.modified else GeneralResearcher().get_info(file)['modified'],
                'last_printed': core_properties.last_printed.strftime('%Y-%m-%d %H:%M:%S') if core_properties.last_printed else '-',
                'category': core_properties.category,
                'pages': len(doc.paragraphs),
                'format': self.get_format(section.page_width / 36000, section.page_height / 36000), #_EMUS_PER_MM = 36000 https://python-docx.readthedocs.io/en/latest/_modules/docx/shared.html
                'orientation': self.get_orientation(section)}
            
class PDFResearcher:
    
    PDF_PAGE_SIZES = {
        'A0': (2384, 3370),
        'A1': (1684, 2384),
        'A2': (1191, 1684),
        'A3': (842, 1191),
        'A4': (595, 842),
        'A5': (420, 595)
    }
    
    def get_format(self, width, height):
        for format, (w, h) in self.PDF_PAGE_SIZES.items():
            if abs(width - w) < 10 and abs(height - h) < 10:
                return format
        return f'{int(width)}x{int(height)}'
    
    def get_orientation(self, width, height):
        return 'landscape' if width > height else 'portrait'
    
    def get_info(self, file):
        if not file.lower().endswith('.pdf'):
            return {}
        else:
            pdf = PdfReader(file)
            page = pdf.pages[0].mediabox
            width, height = page.width, page.height
            return {
                'pages': len(pdf.pages),
                'format': self.get_format(width, height),
                'orientation': self.get_orientation(width, height),
                'metadata': pdf.metadata
            }
    
class ExcelResearcher:
    def get_info(self, file):
        if not file.lower().endswith('.xlsx'):
            return {}
        else:
            wb = load_workbook(file)
            return {
                'sheets': len(wb.sheetnames),
                'author': wb.properties.creator
            }

class AudioResearcher: 
    def get_info(self, file):
        if not file.lower().endswith('.mp3'):
            return {}
        else:
            audio = MP3(file)
            return {
                'duration': round(audio.info.length / 60, 3), #min
                'bitrate': round(audio.info.bitrate / 1000)  #kbps
                }
    
class VideoResearcher:
    def get_info(self, file):
        if not file.lower().endswith(('.mp4', '.avi', '.mkv')):
            return {}
        else:
            cap = cv2.VideoCapture(file)
            width = int(cap.get(cv2.CAP_PROP_FRAME_WIDTH))
            height = int(cap.get(cv2.CAP_PROP_FRAME_HEIGHT))
            fps = cap.get(cv2.CAP_PROP_FPS)
            duration = cap.get(cv2.CAP_PROP_FRAME_COUNT) / fps
            cap.release()
            return {
                'width': width,
                'height': height,
                'fps': round(fps, 3),
                'duration': round(duration / 60, 3) #в минутах
                }
    
class Saver:
    def __init__(self, output_file):
        self.output_file = output_file

    def save_scv(self, data):
        df = pd.DataFrame(data)
        df = df.astype(object)
        df.fillna('-', inplace=True)
        df.to_csv(self.output_file, index=False)
        
class Displayer:
    def __init__(self, output_file):
        self.output_file = output_file

    def display(self, text_widget):
        df = pd.read_csv(self.output_file)
        extension_counts = df['extension'].value_counts()

        total = len(df)
        extension_summary = "\n".join([f"{extension}: {count}" for extension, count in extension_counts.items()])

        summary = f"\nОбщее количество файлов: {total}\n\nКоличество файлов по расширениям:\n{extension_summary}"
        
        text_widget.delete(1.0, tk.END)
        text_widget.insert(tk.END, summary)
        
    def display_top_size_files(self, data, results_area):
        top_files = sorted(data, key=lambda x: x['size'], reverse=True)[:10]

        if top_files:
            top_files_str = "\n\nТоп 10 файлов по размеру:\n\n"
            for idx, file_info in enumerate(top_files, start=1):
                size_formatted = humanize.naturalsize(file_info['size'], binary=True)
                top_files_str += f"{idx}. {file_info['file_path']} - {size_formatted}\n"
            
            results_area.insert(tk.END, top_files_str)
