import pandas as pd
from os.path import join, getsize, getctime, getmtime
from tqdm import tqdm
from datetime import datetime
from PyPDF2 import PdfReader
from PIL import Image
from docx import Document
from openpyxl import load_workbook
from mutagen.mp3 import MP3
import cv2
import logging
import os
from collect.database_saver import DatabaseSaver

class Logger:
    """Logger class to handle logging of errors and exceptions. Logs messages
    with a timestamp to a specified log file.

    Attributes:
        logger: The logger object used to log messages.
    """
    def __init__(self, log_file='stakanov.log'):
        """Initialize the logger.

        Args:
            log_file (str): The path of the log file (default: 'stakanov.log').
        """
        logging.basicConfig(
            filename=log_file,
            format='%(asctime)s - %(levelname)s - %(message)s',
            datefmt='%Y-%m-%d %H:%M:%S'
        )
        self.logger = logging.getLogger()

    def log_exception(self, message):
        """Logs an exception message to the log file."""
        self.logger.exception(message)

class Guide:
    """A class that helps to guide through all the files in a given directory.

    Attributes:
        path (str): The path to the directory to scan for files.
    """
    def __init__(self, path):
        """Initialize the guide.

        Args:
            path (str): The path to the directory to scan for files.
        """
        self.path = path
        
    def files(self):
        """Generator that yields each file in the specified directory and its
        subdirectories."""
        for root, dirs, files in os.walk(self.path):
            for f in files:
                yield join(root, f)
                
class Indiana:
    """Main class for scanning files and gathering their information.

    Attributes:
        logger (Logger): The logger instance to log exceptions.
        guide (Guide): The Guide instance to traverse files in the specified path.
        researchers (list): List of researchers to gather file information.
        output_file (str): The output file where results will be saved.
        data (list): A list to store information about processed files.
        csv_saver (Saver): The instance responsible for saving results to a CSV file.
        viewer (Displayer): The instance responsible for displaying results.
    """
    def  __init__(self, path, output_file):
        """Initializes the Indiana class with the directory path and output
        file name.

        Args:
            path (str): The path to scan for files.
            output_file (str): The name of the output file to save results.
        """
        self.logger = Logger()
        self.guide = Guide(path)
        self.path = path
        self.researchers = [
            GeneralResearcher(), PDFResearcher(), ImageResearcher(), 
            DocResearcher(), ExcelResearcher(), AudioResearcher(), 
            VideoResearcher()
        ]
        self.output_file = output_file
        self.data = []
        self.csv_saver = Saver(output_file)
        self.viewer = Displayer(output_file)
        self.db_saver = DatabaseSaver(self.logger)

    def find_loot(self, progress_callback=None):
        """Scans all files in the specified directory, gathering information
        using different researchers.

        Args:
            progress_callback (function, optional): A callback function to update progress.
        """
        total_files = sum(1 for _ in self.guide.files())  
        processed_files = 0  
    
        if progress_callback:
            progress_callback(processed_files, total_files)

        for file in tqdm(self.guide.files(), total=total_files):
            file_info = {}
            for researcher in self.researchers:
                try:
                    file_info.update(researcher.get_info(file))
                except Exception as e:
                    self.logger.log_exception(f"Ошибка при обработке файла {file}: {str(e)}")

            if file_info:
                file_info['file_path'] = file
                self.data.append(file_info)

            processed_files += 1
            if progress_callback:
                progress_callback(processed_files, total_files)

                
    def save_results(self):
        """Saves data into the output CSV file."""
        self.csv_saver.save_scv(self.data)
    
    def display_results(self):
        """Displays the results in the viewer."""
        self.viewer.display()
        
    def save_to_db(self, run_id):
        """Saves the results in the database."""
        self.db_saver.save_to_db(self.data, run_id)

class GeneralResearcher:
    """A researcher that gathers general information about a file, including
    size, name, extension, and timestamps."""
    def get_info(self, file):
        """Gathers general information about a file.

        Args:
            file (str): The path to the file.

        Returns:
            dict: A dictionary containing general file information (name, extension, size, created, modified).
        """
        size = getsize(file)
        name = os.path.splitext(os.path.basename(file))[0]
        extension = os.path.splitext(file)[1].replace('.', '').upper()
        return {
            'name': name,
            'extension': extension,
            'size': size,
            'created': datetime.fromtimestamp(getctime(file)).strftime('%Y-%m-%d %H:%M:%S'),
            'modified': datetime.fromtimestamp(getmtime(file)).strftime('%Y-%m-%d %H:%M:%S')
        }
        
    
class ImageResearcher:
    """A researcher that gathers information about image files."""
    def get_info(self, file):
        """Gathers information about an image file.

        Args:
            file (str): The path to the image file.

        Returns:
            dict: A dictionary containing image information (width, height, area, dpi, exif).
        """
        if not file.lower().endswith(('.png', '.jpg', '.jpeg', '.gif')):
            return {}
        else:
            img = cv2.imread(file)
            height, width = img.shape[:2]
            img_pil = Image.open(file)
            dpi = img_pil.info.get('dpi', (None, None))
            if dpi != (None, None):
                dpi = (round(dpi[0]))
            exif_data = img_pil._getexif()
            
            return {
                'width': width,
                'height': height,
                'area': width * height, 
                'dpi': dpi if dpi != (None, None) else None,
                'exif': exif_data if exif_data else 'No EXIF data'
            }
    
class DocResearcher:
    """A researcher that gathers information about doc files."""
    DOC_PAGE_SIZES = {
        'A0': (841, 1189),
        'A1': (594, 841),
        'A2': (420, 594),
        'A3': (297, 420),
        'A4': (210, 297),
        'A5': (148, 210),
        'A6': (105, 148)
    }
    
    def get_format(self, width, height):
        """Determines the standard page format of a document based on its
        dimensions.

        Args:
            width (float): The width of the document in millimeters.
            height (float): The height of the document in millimeters.

        Returns:
            str: The name of the standard page format (e.g., 'A4') or custom dimensions as a string.
        """
        for format, (w, h) in self.DOC_PAGE_SIZES.items():
            if abs(width - w) < 10 and abs(height - h) < 10:
                return format
        return f'{int(width)}x{int(height)}'
    
    def get_orientation(self, section):
        """Determines the orientation of a document section.

        Args:
            section (docx.section.Section): A section object from the document.

        Returns:
            str: 'landscape' if the section is in landscape orientation, otherwise 'portrait'.
        """
        return 'landscape' if section.orientation == 1 else 'portrait'
    
    def get_info(self, file):
        """Extracts metadata and format details from file.

        Args:
            file (str): Path to the .docx file.

        Returns:
            dict: A dictionary containing metadata and format information, including:
                - author (str): Author of the document.
                - created (str): Creation timestamp of the document.
                - modified (str): Last modification timestamp.
                - last_printed (str): Timestamp of the last print event.
                - category (str): Document category.
                - pages (int): Estimated number of pages (count of paragraphs).
                - format (str): Page format (e.g., 'A4').
                - orientation (str): Page orientation ('landscape' or 'portrait').
        """
        if not file.lower().endswith('.docx'):
            return {}
        else:
            doc = Document(file)
            core_properties = doc.core_properties
            section = doc.sections[0]
            return {
                'author': core_properties.author,
                'created': core_properties.created.strftime('%Y-%m-%d %H:%M:%S') if core_properties.created else GeneralResearcher().get_info(file)['created'],
                'modified': core_properties.modified.strftime('%Y-%m-%d %H:%M:%S') if core_properties.modified else GeneralResearcher().get_info(file)['modified'],
                'last_printed': core_properties.last_printed.strftime('%Y-%m-%d %H:%M:%S') if core_properties.last_printed else '-',
                'category': core_properties.category,
                'pages': len(doc.paragraphs),
                'format': self.get_format(section.page_width / 36000, section.page_height / 36000), #_EMUS_PER_MM = 36000 https://python-docx.readthedocs.io/en/latest/_modules/docx/shared.html
                'orientation': self.get_orientation(section)}
            
class PDFResearcher:
    """A researcher that gathers information about pdf files."""
    PDF_PAGE_SIZES = {
        'A0': (2384, 3370),
        'A1': (1684, 2384),
        'A2': (1191, 1684),
        'A3': (842, 1191),
        'A4': (595, 842),
        'A5': (420, 595)
    }
    
    def get_format(self, width, height):
        """Determines the standard page format of a pdf page based on its
        dimensions.

        Args:
            width (float): The width of the pdf page in points.
            height (float): The height of the pdf page in points.

        Returns:
            str: The name of the standard page format (e.g., 'A4') or custom dimensions as a string.
        """
        for format, (w, h) in self.PDF_PAGE_SIZES.items():
            if abs(width - w) < 10 and abs(height - h) < 10:
                return format
        return f'{int(width)}x{int(height)}'
    
    def get_orientation(self, width, height):
        """Determines the orientation of a pdf page.

        Args:
            width (float): The width of the pdf page in points.
            height (float): The height of the pdf page in points.

        Returns:
            str: 'landscape' if the page is wider than it is tall, otherwise 'portrait'.
        """
        return 'landscape' if width > height else 'portrait'
    
    def get_info(self, file):
        """Extracts metadata and format details from file.

        Args:
            file (str): Path to the pdf file.

        Returns:
            dict: A dictionary containing metadata and format information, including:
                - pages (int): Total number of pages in the pdf.
                - format (str): Page format (e.g., 'A4').
                - orientation (str): Orientation of the first page ('landscape' or 'portrait').
                - metadata (dict): Metadata from the pdf file (e.g., title, author, etc.).
        """
        if not file.lower().endswith('.pdf'):
            return {}
        else:
            pdf = PdfReader(file)
            page = pdf.pages[0].mediabox
            width, height = page.width, page.height
            return {
                'pages': len(pdf.pages),
                'format': self.get_format(width, height),
                'orientation': self.get_orientation(width, height),
                'metadata': pdf.metadata
            }
    
class ExcelResearcher:
    """A researcher that gathers information about excel files."""
    def get_info(self, file):
        """Extracts information from file.

        Args:
            file (str): Path to the excel file.

        Returns:
            dict: A dictionary containing information, including:
                - sheets (int): Total number of sheets in the workbook.
                - author (str): Author of the workbook.
        """
        if not file.lower().endswith('.xlsx'):
            return {}
        else:
            wb = load_workbook(file)
            return {
                'sheets': len(wb.sheetnames),
                'author': wb.properties.creator
            }

class AudioResearcher: 
    """A researcher class for extracting metadata from audio files."""
    def get_info(self, file):
        """Extracts metadata from an MP3 file.

        Args:
            file (str): Path to the MP3 file.

        Returns:
            dict: A dictionary containing audio metadata, including:
                - duration (float): Duration of the audio file in minutes.
                - bitrate (int): Bitrate of the audio file in kbps.
        """
        if not file.lower().endswith('.mp3'):
            return {}
        else:
            audio = MP3(file)
            return {
                'duration': round(audio.info.length / 60, 3), #min
                'bitrate': round(audio.info.bitrate / 1000)  #kbps
                }
    
class VideoResearcher:
    """A researcher class for extracting metadata from video files."""
    def get_info(self, file):
        """Extracts metadata from a video file.

        Args:
            file (str): Path to the video file.

        Returns:
            dict: A dictionary containing video metadata, including:
                - width (int): Width of the video in pixels.
                - height (int): Height of the video in pixels.
                - fps (float): Frames per second of the video.
                - duration (float): Duration of the video in minutes.
        """
        if not file.lower().endswith(('.mp4', '.avi', '.mkv')):
            return {}
        else:
            cap = cv2.VideoCapture(file)
            width = int(cap.get(cv2.CAP_PROP_FRAME_WIDTH))
            height = int(cap.get(cv2.CAP_PROP_FRAME_HEIGHT))
            fps = cap.get(cv2.CAP_PROP_FPS)
            duration = cap.get(cv2.CAP_PROP_FRAME_COUNT) / fps
            cap.release()
            return {
                'width': width,
                'height': height,
                'fps': round(fps, 3),
                'duration': round(duration / 60, 3) #в минутах
                }
    
class Saver:
    """Сlass for saving data to a CSV file."""
    def __init__(self, output_file):
        """Initializes the Saver with a specified output file.

        Args:
            output_file (str): Path to the output CSV file.
        """
        self.output_file = output_file

    def save_scv(self, data):
        """Saves data to a CSV file.

        Args:
            data (list): A list of dictionaries containing the data to be saved.
        """
        df = pd.DataFrame(data)
        df = df.astype(object)
        df.fillna('-', inplace=True)
        df.to_csv(self.output_file, index=False)